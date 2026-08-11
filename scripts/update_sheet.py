from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

from common import ROOT, load_settings, next_project_id, normalize_design_type, now_iso, read_json, write_json


HEADERS = [
    "Project ID",
    "Status",
    "Customer Name",
    "Company Name",
    "Project Name",
    "Design Type",
    "Address",
    "Started Date",
    "Invoice#",
    "Customer Email",
    "Last Activity",
    "Notes",
    "Evidence",
]


def ensure_headers(ws):
    current = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if "Project ID" not in current:
        ws.insert_cols(1)
        ws.cell(1, 1).value = "Project ID"
    current = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    project_name_col = current.index("Project Name") + 1 if "Project Name" in current else 5
    if "Design Type" not in current:
        ws.insert_cols(project_name_col + 1)
        ws.cell(1, project_name_col + 1).value = "Design Type"
    for col, header in enumerate([ws.cell(1, col).value for col in range(1, ws.max_column + 1)], start=1):
        if header is None and col <= len(HEADERS):
            ws.cell(1, col).value = HEADERS[col - 1]


def header_map(ws) -> dict[str, int]:
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def append_note(existing: str | None, note: str) -> str:
    existing = existing or ""
    if note in existing:
        return existing
    return (existing + "\n" + note).strip()


def main() -> None:
    settings = load_settings()
    cache = ROOT / settings["cache_dir"]
    groups = read_json(cache / "classified_orders.json", {"groups": []})["groups"]
    wb = load_workbook(ROOT / settings["orders"]["workbook"])
    ws = wb[settings["orders"]["sheet"]]
    ensure_headers(ws)
    cols = header_map(ws)
    existing_ids = [str(ws.cell(row, cols["Project ID"]).value or "").strip() for row in range(2, ws.max_row + 1)]
    project_year = settings.get("project_id", {}).get("year", datetime.now().year)
    updates = []

    for group in groups:
        confidence = group.get("status_confidence", 0)
        latest = group["latest_message"]
        note = f"{latest.get('date')}: {latest.get('subject')} [{latest.get('id')}]"
        row_idx = group.get("matched_order_row")
        if row_idx:
            if not ws.cell(row_idx, cols["Project ID"]).value:
                new_project_id = next_project_id(existing_ids, project_year)
                ws.cell(row_idx, cols["Project ID"]).value = new_project_id
                existing_ids.append(new_project_id)
            old_status = ws.cell(row_idx, cols["Status"]).value
            if confidence >= settings["minimum_confidence_for_auto_update"] and old_status != group["suggested_status"]:
                ws.cell(row_idx, cols["Status"]).value = group["suggested_status"]
            if not ws.cell(row_idx, cols["Design Type"]).value:
                ws.cell(row_idx, cols["Design Type"]).value = group.get("design_type") or normalize_design_type(group.get("project", ""))
            ws.cell(row_idx, cols["Last Activity"]).value = latest.get("date")
            ws.cell(row_idx, cols["Notes"]).value = append_note(ws.cell(row_idx, cols["Notes"]).value, note)
            ws.cell(row_idx, cols["Evidence"]).value = append_note(ws.cell(row_idx, cols["Evidence"]).value, latest.get("snippet", ""))
            updates.append({"type": "updated", "project_id": ws.cell(row_idx, cols["Project ID"]).value, "row": row_idx, "status": group["suggested_status"], "confidence": confidence})
        elif confidence >= settings["minimum_confidence_for_auto_update"]:
            row_idx = ws.max_row + 1
            project_id = next_project_id(existing_ids, project_year)
            existing_ids.append(project_id)
            ws.cell(row_idx, cols["Project ID"]).value = project_id
            ws.cell(row_idx, cols["Status"]).value = group["suggested_status"]
            ws.cell(row_idx, cols["Project Name"]).value = group["project"]
            ws.cell(row_idx, cols["Design Type"]).value = group.get("design_type") or normalize_design_type(group.get("project", ""))
            ws.cell(row_idx, cols["Started Date"]).value = datetime.now().strftime("%Y-%m-%d")
            ws.cell(row_idx, cols["Customer Email"]).value = group["customer_email"]
            ws.cell(row_idx, cols["Last Activity"]).value = latest.get("date")
            ws.cell(row_idx, cols["Notes"]).value = note
            ws.cell(row_idx, cols["Evidence"]).value = latest.get("snippet", "")
            updates.append({"type": "created", "project_id": project_id, "row": row_idx, "status": group["suggested_status"], "confidence": confidence})

    wb.save(ROOT / settings["orders"]["workbook"])
    write_json(cache / "updates.json", {"updated_at": now_iso(), "updates": updates})


if __name__ == "__main__":
    main()
