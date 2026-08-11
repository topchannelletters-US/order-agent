from __future__ import annotations

from openpyxl import load_workbook

from common import ROOT, load_settings, norm, read_json, write_json


def row_dict(headers, row):
    return {headers[i]: row[i].value if i < len(row) else None for i in range(len(headers))}


def main() -> None:
    settings = load_settings()
    cache = ROOT / settings["cache_dir"]
    groups = read_json(cache / "grouped_threads.json", {"groups": []})["groups"]
    wb = load_workbook(ROOT / settings["orders"]["workbook"])
    ws = wb[settings["orders"]["sheet"]]
    headers = [c.value for c in ws[1]]
    orders = []
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        data = row_dict(headers, row)
        haystack = norm(" ".join(str(data.get(k) or "") for k in ["Project ID", "Customer Email", "Customer Name", "Company Name", "Project Name", "Design Type", "Address", "Invoice#"]))
        orders.append({"row": idx, "data": data, "haystack": haystack})

    matched = []
    for group in groups:
        text = group.get("match_text") or norm(" ".join([group.get("project_id", ""), group["customer_email"], group["project"], group.get("design_type", "")]))
        best = None
        best_score = 0
        best_reasons = []
        for order in orders:
            score = 0
            reasons = []
            project_id = norm(str(order["data"].get("Project ID") or ""))
            incoming_project_id = norm(group.get("project_id", ""))
            if project_id and incoming_project_id and project_id == incoming_project_id:
                score += 100
                reasons.append("project_id")
            email = norm(str(order["data"].get("Customer Email") or ""))
            project = norm(str(order["data"].get("Project Name") or ""))
            address = norm(str(order["data"].get("Address") or ""))
            design_type = norm(str(order["data"].get("Design Type") or ""))
            if project and project in text:
                score += 35
                reasons.append("project_name")
            if address and address in text:
                score += 25
                reasons.append("address")
            if email and email in text:
                score += 15
                reasons.append("email")
            if design_type and design_type in text and design_type != "unknown review":
                score += 5
                reasons.append("design_type")
            if score > best_score:
                best = order
                best_score = score
                best_reasons = reasons
        group["matched_order_row"] = best["row"] if best and best_score >= 35 else None
        group["match_confidence"] = best_score
        group["match_reasons"] = best_reasons
        matched.append(group)
    write_json(cache / "matched_orders.json", {"groups": matched})


if __name__ == "__main__":
    main()
